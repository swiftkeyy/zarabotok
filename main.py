#!/usr/bin/env python3
"""
💼 Бот "Календарь Заработка" — SQLite/PostgreSQL версия
Поддержка Railway (PostgreSQL) и bothost.ru (SQLite)
"""

import asyncio
import json
import logging
import os
from datetime import datetime, timedelta
from io import BytesIO
from typing import List

from aiogram import Bot, Dispatcher, F
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.filters import Command, CommandStart, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton,
    ReplyKeyboardMarkup, KeyboardButton, WebAppInfo
)
from aiogram.utils.keyboard import InlineKeyboardBuilder
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from openpyxl import Workbook

# ================== НАСТРОЙКИ ==================
BOT_TOKEN = os.getenv("BOT_TOKEN", "YOUR_BOT_TOKEN_HERE")
ADMIN_IDS = [int(x) for x in os.getenv("ADMIN_IDS", "123456789").split(",")]
WEBAPP_URL = os.getenv("WEBAPP_URL", "https://твой-бот.bothost.ru/miniapp/work-earn-miniapp.html")

# Проверка токена
if not BOT_TOKEN or BOT_TOKEN == "YOUR_BOT_TOKEN_HERE":
    raise ValueError("❌ BOT_TOKEN не установлен! Укажи переменную окружения BOT_TOKEN")

# Database configuration
DATABASE_URL = os.getenv("DATABASE_URL")  # Railway PostgreSQL
USE_POSTGRES = DATABASE_URL is not None

if USE_POSTGRES:
    # Railway: PostgreSQL
    import asyncpg
    DB_POOL = None
    logger_msg = "🐘 PostgreSQL (Railway)"
else:
    # bothost.ru: SQLite
    import aiosqlite
    DB_PATH = os.getenv("DB_PATH", "/app/data/bot_data.db")
    try:
        os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    except:
        pass
    logger_msg = f"📦 SQLite ({DB_PATH})"

class Broadcast(StatesGroup):
    waiting_for_text = State()
    waiting_for_confirmation = State()

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
logger = logging.getLogger(__name__)

bot = Bot(token=BOT_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher(storage=MemoryStorage())
scheduler = AsyncIOScheduler()

# ================== БАЗА (SQLite/PostgreSQL) ==================
async def init_db():
    global DB_POOL
    
    if USE_POSTGRES:
        # PostgreSQL (Railway)
        DB_POOL = await asyncpg.create_pool(DATABASE_URL, min_size=1, max_size=10)
        
        async with DB_POOL.acquire() as conn:
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    user_id BIGINT PRIMARY KEY,
                    username TEXT,
                    first_name TEXT,
                    joined_at TIMESTAMP,
                    last_active TIMESTAMP
                )
            """)
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS stats (
                    id SERIAL PRIMARY KEY,
                    user_id BIGINT,
                    month TEXT,
                    work_days INTEGER,
                    earnings INTEGER,
                    rate INTEGER,
                    passive_rate INTEGER,
                    saved_at TIMESTAMP
                )
            """)
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS forced_channels (
                    id SERIAL PRIMARY KEY,
                    channel_id BIGINT UNIQUE,
                    channel_username TEXT,
                    title TEXT,
                    added_by BIGINT,
                    added_at TIMESTAMP
                )
            """)
        logger.info(f"✅ {logger_msg} база готова")
    else:
        # SQLite (bothost.ru)
        async with aiosqlite.connect(DB_PATH) as db:
            await db.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    user_id INTEGER PRIMARY KEY,
                    username TEXT,
                    first_name TEXT,
                    joined_at TEXT,
                    last_active TEXT
                )
            """)
            await db.execute("""
                CREATE TABLE IF NOT EXISTS stats (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER,
                    month TEXT,
                    work_days INTEGER,
                    earnings INTEGER,
                    rate INTEGER,
                    passive_rate INTEGER,
                    saved_at TEXT
                )
            """)
            await db.execute("""
                CREATE TABLE IF NOT EXISTS forced_channels (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    channel_id INTEGER UNIQUE,
                    channel_username TEXT,
                    title TEXT,
                    added_by INTEGER,
                    added_at TEXT
                )
            """)
            await db.commit()
        logger.info(f"✅ {logger_msg} база готова")

async def add_or_update_user(user_id: int, username: str, first_name: str) -> bool:
    now = datetime.now()
    
    if USE_POSTGRES:
        async with DB_POOL.acquire() as conn:
            existing = await conn.fetchrow("SELECT joined_at FROM users WHERE user_id = $1", user_id)
            is_new = existing is None
            
            if is_new:
                await conn.execute("""
                    INSERT INTO users (user_id, username, first_name, joined_at, last_active)
                    VALUES ($1, $2, $3, $4, $5)
                """, user_id, username, first_name, now, now)
            else:
                await conn.execute("""
                    UPDATE users SET username = $1, first_name = $2, last_active = $3
                    WHERE user_id = $4
                """, username, first_name, now, user_id)
            return is_new
    else:
        now_str = now.isoformat()
        async with aiosqlite.connect(DB_PATH) as db:
            cursor = await db.execute("SELECT joined_at FROM users WHERE user_id = ?", (user_id,))
            is_new = (await cursor.fetchone()) is None
            
            await db.execute("""
                INSERT OR REPLACE INTO users (user_id, username, first_name, joined_at, last_active)
                VALUES (?, ?, ?, COALESCE((SELECT joined_at FROM users WHERE user_id = ?), ?), ?)
            """, (user_id, username, first_name, user_id, now_str, now_str))
            await db.commit()
            return is_new

async def notify_admins_new_user(user: dict):
    for admin_id in ADMIN_IDS:
        try:
            await bot.send_message(admin_id, f"🆕 Новый пользователь: <b>{user['first_name']}</b>")
        except: pass

async def save_stats_from_miniapp(user_id: int, data: dict):
    now = datetime.now()
    
    if USE_POSTGRES:
        async with DB_POOL.acquire() as conn:
            await conn.execute("""
                INSERT INTO stats (user_id, month, work_days, earnings, rate, passive_rate, saved_at)
                VALUES ($1, $2, $3, $4, $5, $6, $7)
            """, user_id, data.get("month"), data.get("work_days", 0), data.get("earnings", 0),
                  data.get("rate", 0), data.get("passive_rate", 0), now)
    else:
        async with aiosqlite.connect(DB_PATH) as db:
            await db.execute("""
                INSERT INTO stats (user_id, month, work_days, earnings, rate, passive_rate, saved_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (user_id, data.get("month"), data.get("work_days", 0), data.get("earnings", 0),
                  data.get("rate", 0), data.get("passive_rate", 0), now.isoformat()))
            await db.commit()

    # Уведомления о достижениях
    earnings = data.get("earnings", 0)
    work_days = data.get("work_days", 0)
    month = data.get("month", "")

    if earnings >= 200000:
        for admin_id in ADMIN_IDS:
            try:
                await bot.send_message(admin_id, f"🏆 КРУТОЕ ДОСТИЖЕНИЕ! Пользователь {user_id} заработал {earnings:,} ₽ в {month}!")
            except: pass

    if work_days >= 26:
        for admin_id in ADMIN_IDS:
            try:
                await bot.send_message(admin_id, f"💪 ЖЕЛЕЗНЫЙ ЧЕЛОВЕК! Пользователь {user_id} отработал {work_days} дней в {month}!")
            except: pass

async def get_total_stats():
    if USE_POSTGRES:
        async with DB_POOL.acquire() as conn:
            row = await conn.fetchrow("""
                SELECT COUNT(DISTINCT user_id), COUNT(*), COALESCE(SUM(earnings),0), COALESCE(AVG(earnings),0) 
                FROM stats
            """)
            return {"users": row[0] or 0, "total_saves": row[1] or 0, "total_earnings": row[2] or 0, "avg_earnings": round(row[3] or 0)}
    else:
        async with aiosqlite.connect(DB_PATH) as db:
            cursor = await db.execute("SELECT COUNT(DISTINCT user_id), COUNT(*), COALESCE(SUM(earnings),0), COALESCE(AVG(earnings),0) FROM stats")
            row = await cursor.fetchone()
            return {"users": row[0] or 0, "total_saves": row[1] or 0, "total_earnings": row[2] or 0, "avg_earnings": round(row[3] or 0)}

async def get_users_without_recent_stats(days: int = 30):
    cutoff = datetime.now() - timedelta(days=days)
    
    if USE_POSTGRES:
        async with DB_POOL.acquire() as conn:
            rows = await conn.fetch("""
                SELECT u.user_id, u.first_name 
                FROM users u LEFT JOIN stats s ON u.user_id = s.user_id AND s.saved_at > $1
                WHERE s.id IS NULL
            """, cutoff)
            return [(r['user_id'], r['first_name']) for r in rows]
    else:
        cutoff_str = cutoff.isoformat()
        async with aiosqlite.connect(DB_PATH) as db:
            cursor = await db.execute("""
                SELECT u.user_id, u.first_name 
                FROM users u LEFT JOIN stats s ON u.user_id = s.user_id AND s.saved_at > ?
                WHERE s.id IS NULL
            """, (cutoff_str,))
            return await cursor.fetchall()

async def get_weekly_report_data():
    week_ago = datetime.now() - timedelta(days=7)
    
    if USE_POSTGRES:
        async with DB_POOL.acquire() as conn:
            new_users = await conn.fetchval("SELECT COUNT(*) FROM users WHERE joined_at > $1", week_ago)
            stats_row = await conn.fetchrow("SELECT COUNT(*), COALESCE(SUM(earnings), 0) FROM stats WHERE saved_at > $1", week_ago)
            top = await conn.fetch("""
                SELECT u.first_name, COALESCE(SUM(s.earnings), 0) as total
                FROM users u LEFT JOIN stats s ON u.user_id = s.user_id
                GROUP BY u.user_id, u.first_name ORDER BY total DESC LIMIT 5
            """)
            return {
                "new_users": new_users or 0,
                "saves_this_week": stats_row[0] or 0,
                "earnings_this_week": stats_row[1] or 0,
                "top_users": [(r['first_name'], r['total']) for r in top]
            }
    else:
        week_ago_str = week_ago.isoformat()
        async with aiosqlite.connect(DB_PATH) as db:
            new_users = await (await db.execute("SELECT COUNT(*) FROM users WHERE joined_at > ?", (week_ago_str,))).fetchone()
            row = await (await db.execute("SELECT COUNT(*), COALESCE(SUM(earnings), 0) FROM stats WHERE saved_at > ?", (week_ago_str,))).fetchone()
            top = await (await db.execute("""
                SELECT u.first_name, COALESCE(SUM(s.earnings), 0) as total
                FROM users u LEFT JOIN stats s ON u.user_id = s.user_id
                GROUP BY u.user_id ORDER BY total DESC LIMIT 5
            """)).fetchall()
            return {
                "new_users": new_users[0] or 0,
                "saves_this_week": row[0] or 0,
                "earnings_this_week": row[1] or 0,
                "top_users": top
            }

# ================== ОБЯЗАТЕЛЬНАЯ ПОДПИСКА ==================
async def get_forced_channels():
    if USE_POSTGRES:
        async with DB_POOL.acquire() as conn:
            rows = await conn.fetch("SELECT channel_id, channel_username, title FROM forced_channels")
            return [(r['channel_id'], r['channel_username'], r['title']) for r in rows]
    else:
        async with aiosqlite.connect(DB_PATH) as db:
            cursor = await db.execute("SELECT channel_id, channel_username, title FROM forced_channels")
            return await cursor.fetchall()

async def is_user_subscribed(user_id: int, channel_id: int) -> bool:
    try:
        member = await bot.get_chat_member(chat_id=channel_id, user_id=user_id)
        return member.status in ["member", "administrator", "creator"]
    except:
        return False

async def check_forced_subscription(user_id: int) -> list:
    channels = await get_forced_channels()
    not_subscribed = []
    for ch in channels:
        if not await is_user_subscribed(user_id, ch[0]):
            not_subscribed.append(ch)
    return not_subscribed

async def add_forced_channel(channel_id: int, username: str, title: str, added_by: int):
    now = datetime.now()
    
    if USE_POSTGRES:
        async with DB_POOL.acquire() as conn:
            await conn.execute("""
                INSERT INTO forced_channels (channel_id, channel_username, title, added_by, added_at)
                VALUES ($1, $2, $3, $4, $5)
                ON CONFLICT (channel_id) DO UPDATE SET
                    channel_username = $2, title = $3, added_by = $4, added_at = $5
            """, channel_id, username, title, added_by, now)
    else:
        async with aiosqlite.connect(DB_PATH) as db:
            await db.execute("""
                INSERT OR REPLACE INTO forced_channels (channel_id, channel_username, title, added_by, added_at)
                VALUES (?, ?, ?, ?, ?)
            """, (channel_id, username, title, added_by, now.isoformat()))
            await db.commit()

async def remove_forced_channel(channel_id: int):
    if USE_POSTGRES:
        async with DB_POOL.acquire() as conn:
            await conn.execute("DELETE FROM forced_channels WHERE channel_id = $1", channel_id)
    else:
        async with aiosqlite.connect(DB_PATH) as db:
            await db.execute("DELETE FROM forced_channels WHERE channel_id = ?", (channel_id,))
            await db.commit()

# ================== НАПОМИНАНИЯ И ОТЧЁТЫ ==================
async def send_personal_reminders():
    inactive = await get_users_without_recent_stats(30)
    if not inactive: return
    
    text = "👋 Привет!\n\nТы не сохранял статистику уже больше 30 дней.\nОткрой Mini App и обнови данные!"
    for user_id, name in inactive:
        try:
            await bot.send_message(user_id, text, reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                InlineKeyboardButton(text="📅 Открыть календарь", web_app=WebAppInfo(url=WEBAPP_URL))
            ]]))
            await asyncio.sleep(0.1)
        except: pass

async def send_weekly_report():
    data = await get_weekly_report_data()
    top_text = "\n".join([f"• <b>{name}</b> — {total:,} ₽" for name, total in data["top_users"]]) or "—"
    report = f"📊 <b>Еженедельный отчёт</b>\n\n🆕 Новых: <b>{data['new_users']}</b>\n📝 Сохранений: <b>{data['saves_this_week']}</b>\n💰 Заработок: <b>{data['earnings_this_week']:,} ₽</b>\n\n🏆 Топ-5:\n{top_text}"
    for admin_id in ADMIN_IDS:
        try: await bot.send_message(admin_id, report)
        except: pass

async def send_monthly_reminder():
    text = "📅 <b>Напоминание!</b> Не забудь сохранить статистику за прошлый месяц!"
    
    if USE_POSTGRES:
        async with DB_POOL.acquire() as conn:
            users = await conn.fetch("SELECT user_id FROM users")
            user_ids = [r['user_id'] for r in users]
    else:
        async with aiosqlite.connect(DB_PATH) as db:
            users = await (await db.execute("SELECT user_id FROM users")).fetchall()
            user_ids = [uid for (uid,) in users]
    
    for uid in user_ids:
        try:
            await bot.send_message(uid, text, reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                InlineKeyboardButton(text="📅 Открыть календарь", web_app=WebAppInfo(url=WEBAPP_URL))
            ]]))
            await asyncio.sleep(0.08)
        except: 
            pass

def setup_scheduler():
    scheduler.add_job(send_monthly_reminder, trigger="cron", day=1, hour=10, minute=0, timezone="Europe/Moscow")
    scheduler.add_job(send_personal_reminders, trigger="cron", day_of_week="sun", hour=18, minute=0, timezone="Europe/Moscow")
    scheduler.add_job(send_weekly_report, trigger="cron", day_of_week="mon", hour=9, minute=0, timezone="Europe/Moscow")
    scheduler.start()
    logger.info("⏰ Планировщик запущен")

# ================== КЛАВИАТУРЫ ==================
def main_menu_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="📅 Открыть календарь", web_app=WebAppInfo(url=WEBAPP_URL))],
                  [KeyboardButton(text="📊 Моя статистика"), KeyboardButton(text="❓ Помощь")]],
        resize_keyboard=True
    )

def admin_menu_keyboard():
    builder = InlineKeyboardBuilder()
    builder.button(text="📊 Статистика", callback_data="admin_stats")
    builder.button(text="👥 Пользователи", callback_data="admin_users")
    builder.button(text="📢 Обязательные каналы", callback_data="admin_channels")
    builder.button(text="📢 Рассылка", callback_data="admin_broadcast")
    builder.button(text="📥 Экспорт Excel", callback_data="admin_export")
    builder.button(text="🔄 Обновить", callback_data="admin_refresh")
    builder.adjust(2)
    return builder.as_markup()

# ================== MIDDLEWARE ДЛЯ ПРОВЕРКИ ПОДПИСКИ ==================
async def check_subscription_middleware(handler, event, data):
    """Проверяет подписку перед выполнением команд (кроме админских и callback)"""
    if isinstance(event, Message):
        # Пропускаем админские команды
        if event.text and event.text.startswith(('/admin', '/addchannel', '/removechannel')):
            return await handler(event, data)
        
        # Пропускаем web_app_data (данные из Mini App)
        if event.web_app_data:
            return await handler(event, data)
        
        # Проверяем подписку
        not_subscribed = await check_forced_subscription(event.from_user.id)
        if not_subscribed:
            text = "📢 <b>Для использования бота подпишись на каналы:</b>\n\n"
            buttons = []
            for ch in not_subscribed:
                text += f"• {ch[2]}\n"
                buttons.append([InlineKeyboardButton(text=f"📢 {ch[2]}", url=f"https://t.me/{ch[1].lstrip('@')}")])
            buttons.append([InlineKeyboardButton(text="✅ Проверить подписку", callback_data="check_subscription")])
            await event.answer(text, reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))
            return
    
    return await handler(event, data)

# Регистрируем middleware
dp.message.middleware(check_subscription_middleware)

# ================== ХЕНДЛЕРЫ ==================
@dp.message(CommandStart())
async def cmd_start(message: Message):
    user = message.from_user
    is_new = await add_or_update_user(user.id, user.username or "", user.first_name or "Гость")
    if is_new:
        await notify_admins_new_user({"user_id": user.id, "first_name": user.first_name})
    
    await message.answer(
        f"👋 <b>Привет, {user.first_name}!</b>\n\n"
        f"Это <b>Календарь Заработка</b> — твой личный трекер доходов.\n\n"
        f"📅 Открой Mini App и начни отслеживать свою статистику!",
        reply_markup=main_menu_keyboard()
    )

@dp.callback_query(F.data == "check_subscription")
async def check_subscription(callback: CallbackQuery):
    not_sub = await check_forced_subscription(callback.from_user.id)
    if not_sub:
        await callback.answer("❌ Подпишись на все каналы!", show_alert=True)
    else:
        await callback.message.delete()
        await callback.message.answer(
            f"✅ <b>Отлично!</b>\n\n"
            f"Теперь ты можешь пользоваться ботом.",
            reply_markup=main_menu_keyboard()
        )
        await callback.answer("✅ Подписка подтверждена!")

@dp.message(F.text == "📊 Моя статистика")
async def my_stats(message: Message):
    await message.answer(
        "📊 <b>Твоя статистика</b>\n\n"
        "Открой Mini App, чтобы посмотреть детальную статистику и добавить новые записи.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="📅 Открыть календарь", web_app=WebAppInfo(url=WEBAPP_URL))
        ]])
    )

@dp.message(F.text == "❓ Помощь")
async def help_cmd(message: Message):
    await message.answer(
        "❓ <b>Как пользоваться ботом:</b>\n\n"
        "1. Открой Mini App через кнопку ниже\n"
        "2. Заполняй статистику каждый месяц\n"
        "3. Отслеживай свой прогресс в календаре\n\n"
        "💡 Бот будет напоминать тебе о сохранении статистики!",
        reply_markup=main_menu_keyboard()
    )

@dp.message(Command("admin"))
async def admin_panel(message: Message):
    if message.from_user.id not in ADMIN_IDS: return await message.answer("🚫 Нет доступа")
    await message.answer("🛠️ <b>Админ-панель</b>", reply_markup=admin_menu_keyboard())

@dp.callback_query(F.data == "admin_stats")
async def admin_stats(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS: return await callback.answer("Нет доступа", show_alert=True)
    stats = await get_total_stats()
    text = f"📊 <b>Статистика</b>\n\n👥 Пользователей: <b>{stats['users']}</b>\n📝 Сохранений: <b>{stats['total_saves']}</b>\n💰 Общий заработок: <b>{stats['total_earnings']:,} ₽</b>"
    await callback.message.edit_text(text, reply_markup=admin_menu_keyboard())
    await callback.answer()

@dp.callback_query(F.data == "admin_users")
async def admin_users(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS: return await callback.answer("Нет доступа", show_alert=True)
    
    if USE_POSTGRES:
        async with DB_POOL.acquire() as conn:
            users = await conn.fetch("SELECT first_name, username, joined_at FROM users ORDER BY joined_at DESC LIMIT 10")
            user_list = [(r['first_name'], r['username'], r['joined_at']) for r in users]
    else:
        async with aiosqlite.connect(DB_PATH) as db:
            user_list = await (await db.execute("SELECT first_name, username, joined_at FROM users ORDER BY joined_at DESC LIMIT 10")).fetchall()
    
    text = "👥 <b>Пользователи:</b>\n" + "\n".join([f"• {u[0]} @{u[1] or '—'}" for u in user_list])
    await callback.message.edit_text(text, reply_markup=admin_menu_keyboard())
    await callback.answer()

@dp.callback_query(F.data == "admin_channels")
async def admin_channels(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS: return await callback.answer("Нет доступа", show_alert=True)
    channels = await get_forced_channels()
    text = "📢 <b>Обязательные каналы:</b>\n" + "\n".join([f"• {ch[2]} (@{ch[1]})" for ch in channels]) if channels else "Нет каналов.\n/addchannel @username"
    await callback.message.edit_text(text, reply_markup=admin_menu_keyboard())
    await callback.answer()

@dp.message(Command("addchannel"))
async def add_channel_cmd(message: Message):
    if message.from_user.id not in ADMIN_IDS: return await message.answer("🚫 Нет доступа")
    args = message.text.split(maxsplit=1)
    if len(args) < 2: return await message.answer("Использование: /addchannel @username")
    username = args[1].strip()
    try:
        chat = await bot.get_chat(username)
        await add_forced_channel(chat.id, username, chat.title, message.from_user.id)
        await message.answer(f"✅ Канал {chat.title} добавлен!")
    except Exception as e:
        await message.answer(f"❌ Ошибка: {e}")

@dp.message(Command("removechannel"))
async def remove_channel_cmd(message: Message):
    if message.from_user.id not in ADMIN_IDS: return await message.answer("🚫 Нет доступа")
    args = message.text.split()
    if len(args) < 2: return await message.answer("Использование: /removechannel ID")
    await remove_forced_channel(int(args[1]))
    await message.answer("✅ Канал удалён.")

@dp.callback_query(F.data == "admin_broadcast")
async def start_broadcast(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in ADMIN_IDS: return await callback.answer("Нет доступа", show_alert=True)
    await callback.message.edit_text("📢 Напиши текст:", reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="❌ Отмена", callback_data="admin_cancel")]]))
    await state.set_state(Broadcast.waiting_for_text)
    await callback.answer()

@dp.message(StateFilter(Broadcast.waiting_for_text))
async def broadcast_text(message: Message, state: FSMContext):
    if message.from_user.id not in ADMIN_IDS: return
    await state.update_data(text=message.text)
    await message.answer(f"📢 <b>Предпросмотр:</b>\n\n{message.text}\n\nОтправить?", reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="✅ Отправить", callback_data="broadcast_confirm")],[InlineKeyboardButton(text="❌ Отмена", callback_data="admin_cancel")]]))
    await state.set_state(Broadcast.waiting_for_confirmation)

@dp.callback_query(F.data == "broadcast_confirm", StateFilter(Broadcast.waiting_for_confirmation))
async def broadcast_confirm(callback: CallbackQuery, state: FSMContext):
    text = (await state.get_data()).get("text", "")
    
    if USE_POSTGRES:
        async with DB_POOL.acquire() as conn:
            users = await conn.fetch("SELECT user_id FROM users")
            user_ids = [r['user_id'] for r in users]
    else:
        async with aiosqlite.connect(DB_PATH) as db:
            users = await (await db.execute("SELECT user_id FROM users")).fetchall()
            user_ids = [uid for (uid,) in users]
    
    sent = 0
    for uid in user_ids:
        try: await bot.send_message(uid, text); sent += 1; await asyncio.sleep(0.05)
        except: pass
    await callback.message.edit_text(f"✅ Отправлено {sent} пользователям", reply_markup=admin_menu_keyboard())
    await state.clear()
    await callback.answer()

@dp.callback_query(F.data == "admin_export")
async def admin_export(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS: return await callback.answer("Нет доступа", show_alert=True)
    await callback.answer("⏳ Генерирую...")
    
    if USE_POSTGRES:
        async with DB_POOL.acquire() as conn:
            users = await conn.fetch("""
                SELECT u.user_id, u.first_name, u.username, u.joined_at, COUNT(s.id), COALESCE(SUM(s.earnings), 0)
                FROM users u LEFT JOIN stats s ON u.user_id = s.user_id GROUP BY u.user_id ORDER BY 6 DESC
            """)
            user_data = [(r['user_id'], r['first_name'], r['username'], r['joined_at'], r['count'], r['coalesce']) for r in users]
    else:
        async with aiosqlite.connect(DB_PATH) as db:
            user_data = await (await db.execute("""
                SELECT u.user_id, u.first_name, u.username, u.joined_at, COUNT(s.id), COALESCE(SUM(s.earnings), 0)
                FROM users u LEFT JOIN stats s ON u.user_id = s.user_id GROUP BY u.user_id ORDER BY 6 DESC
            """)).fetchall()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Статистика"
    for col, h in enumerate(["ID", "Имя", "Username", "Дата", "Сохранений", "Заработок"], 1):
        ws.cell(row=1, column=col, value=h).font = {"bold": True}
    for row_idx, u in enumerate(user_data, 2):
        for col, val in enumerate(u, 1):
            ws.cell(row=row_idx, column=col, value=val)
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    await bot.send_document(callback.from_user.id, document=buffer, filename="stats.xlsx")

@dp.callback_query(F.data == "admin_refresh")
async def admin_refresh(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS: return await callback.answer("Нет доступа", show_alert=True)
    await callback.message.edit_text("🔄 Обновлено!", reply_markup=admin_menu_keyboard())
    await callback.answer()

@dp.message(F.web_app_data)
async def handle_webapp_data(message: Message):
    try:
        data = json.loads(message.web_app_data.data)
        await save_stats_from_miniapp(message.from_user.id, data)
        await message.answer("✅ Статистика сохранена!")
    except Exception as e:
        logger.error(f"web_app_data: {e}")
        await message.answer("❌ Ошибка.")

# ================== ЗАПУСК ==================
async def main():
    global DB_POOL
    
    await init_db()
    setup_scheduler()
    logger.info(f"🚀 Бот запущен на {logger_msg}")
    
    try:
        await dp.start_polling(bot)
    finally:
        if USE_POSTGRES and DB_POOL:
            await DB_POOL.close()
            logger.info("🐘 PostgreSQL pool closed")

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("👋 Бот остановлен")